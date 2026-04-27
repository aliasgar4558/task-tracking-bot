"""Shared task log storage and helpers for CLI and GUI."""

from __future__ import annotations

import json
import shutil
import sys
from datetime import date
from pathlib import Path
from typing import Any

try:
    from tabulate import tabulate as _tabulate  # type: ignore
except Exception:  # pragma: no cover
    _tabulate = None

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Font  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None
    Alignment = None  # type: ignore
    Font = None  # type: ignore

DATA_FILE = Path.home() / ".taskbot" / "task_logs.json"

REPORT_HEADERS = [
    "#",
    "Project",
    "Task Title",
    "Task Description",
    "Challenges / Blockers",
    "Efforts hrs",
]


def _cell_lines(value: Any) -> list[str]:
    if value is None:
        return [""]
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    lines = s.split("\n")
    return lines if lines else [""]


def _render_grid(headers: list[str], rows: list[list[Any]]) -> str:
    cols = len(headers)
    if cols == 0:
        return ""

    norm_rows: list[list[list[str]]] = []
    for r in rows:
        cells = list(r)
        if len(cells) < cols:
            cells.extend([""] * (cols - len(cells)))
        elif len(cells) > cols:
            cells = cells[:cols]
        norm_rows.append([_cell_lines(c) for c in cells])

    widths: list[int] = []
    for i in range(cols):
        w = len(headers[i])
        for r in norm_rows:
            for line in r[i]:
                w = max(w, len(line))
        widths.append(w)

    def line(ch_left: str, ch_mid: str, ch_right: str, fill: str) -> str:
        parts = [fill * (w + 2) for w in widths]
        return ch_left + ch_mid.join(parts) + ch_right

    def render_row(multicells: list[list[str]]) -> list[str]:
        height = max(len(c) for c in multicells) if multicells else 1
        out_lines: list[str] = []
        for h in range(height):
            parts = []
            for i in range(cols):
                text = multicells[i][h] if h < len(multicells[i]) else ""
                parts.append(f" {text.ljust(widths[i])} ")
            out_lines.append("|" + "|".join(parts) + "|")
        return out_lines

    out = [
        line("+", "+", "+", "-"),
        "|" + "|".join([f" {headers[i].ljust(widths[i])} " for i in range(cols)]) + "|",
        line("+", "+", "+", "-"),
    ]
    for r in norm_rows:
        out.extend(render_row(r))
    out.append(line("+", "+", "+", "-"))
    return "\n".join(out)


def _tabulate_grid(headers: list[str], rows: list[list[Any]]) -> str:
    for r in rows:
        for c in r:
            if c is not None and ("\n" in str(c) or "\r" in str(c)):
                return _render_grid(headers, rows)
    if _tabulate is None:
        return _render_grid(headers, rows)
    return _tabulate(rows, headers=headers, tablefmt="grid")


def get_today_date() -> str:
    return date.today().isoformat()


def load_tasks() -> dict[str, Any]:
    if not DATA_FILE.exists():
        return {}
    raw = DATA_FILE.read_text(encoding="utf-8").strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        if not isinstance(data, dict):
            raise ValueError("Root must be a JSON object")
        return data
    except (json.JSONDecodeError, ValueError) as e:
        backup = DATA_FILE.with_suffix(".json.corrupt")
        try:
            backup.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(DATA_FILE, backup)
            print(
                f"Warning: task_logs.json was invalid ({e}). "
                f"A copy was saved to {backup}. Starting with empty logs.",
                file=sys.stderr,
            )
        except OSError:
            print(
                f"Warning: task_logs.json was invalid ({e}). Starting with empty logs.",
                file=sys.stderr,
            )
        try:
            DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
            DATA_FILE.write_text("{}\n", encoding="utf-8")
        except OSError:
            pass
        return {}


def save_tasks(data: dict[str, Any]) -> None:
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp = DATA_FILE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    tmp.replace(DATA_FILE)


def get_next_task_id(tasks_for_today: list[dict[str, Any]]) -> int:
    if not tasks_for_today:
        return 1
    ids = [t.get("id", 0) for t in tasks_for_today if isinstance(t.get("id"), int)]
    return max(ids, default=0) + 1


def normalize_optional(value: str) -> str:
    s = value.strip()
    return s if s else "-"


def validate_efforts(value: str) -> float | None:
    s = value.strip()
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    if v < 0:
        return None
    return v


def format_hours(h: float) -> str:
    if h == int(h):
        return str(int(h))
    return str(round(h, 4)).rstrip("0").rstrip(".")


def tasks_table_rows(tasks: list[dict[str, Any]]) -> tuple[list[list[Any]], float]:
    rows: list[list[Any]] = []
    total = 0.0
    seq = 0
    for t in tasks:
        if not isinstance(t, dict):
            continue
        seq += 1
        e = t.get("efforts_hrs")
        try:
            hrs = float(e) if e is not None else 0.0
        except (TypeError, ValueError):
            hrs = 0.0
        total += hrs
        rows.append(
            [
                seq,
                t.get("project", "-"),
                t.get("task_title", ""),
                t.get("task_description", ""),
                t.get("blockers", ""),
                format_hours(hrs),
            ]
        )
    return rows, total


def sorted_today_tasks(
    data: dict[str, Any],
    *,
    group_by_project: bool = False,
) -> tuple[str, list[dict[str, Any]]] | None:
    today = get_today_date()
    tasks = data.get(today)
    if not tasks or not isinstance(tasks, list):
        return None
    valid = [t for t in tasks if isinstance(t, dict)]
    if not valid:
        return None
    if group_by_project:
        def _key(x: dict[str, Any]) -> tuple[int, str, int]:
            p = str(x.get("project", "-") or "-").strip()
            p_norm = p.lower()
            p_is_blank = 1 if (not p or p == "-") else 0
            tid = x.get("id", 0)
            tid_int = tid if isinstance(tid, int) else 0
            return (p_is_blank, p_norm, tid_int)

        valid.sort(key=_key)
    else:
        valid.sort(key=lambda x: int(x.get("id", 0)) if isinstance(x.get("id"), int) else 0)
    return today, valid


def format_report_table(date_str: str, tasks: list[dict[str, Any]]) -> str:
    rows, total = tasks_table_rows(tasks)
    body = _tabulate_grid(REPORT_HEADERS, rows)
    return f"Date: {date_str}\n\n{body}\n\nTotal Efforts: {format_hours(total)} hrs"


def append_task(
    project: str,
    title: str,
    description: str,
    blockers: str,
    efforts_hrs: float,
) -> tuple[bool, str]:
    project = normalize_optional(project)
    title = title.strip()
    if not title:
        return False, "Task title is required."

    data = load_tasks()
    today = get_today_date()
    if today not in data:
        data[today] = []
    bucket = data[today]
    if not isinstance(bucket, list):
        return False, "Today's entry in logs is corrupted. Cannot add tasks."

    entry = {
        "id": get_next_task_id(bucket),
        "project": project,
        "task_title": title,
        "task_description": normalize_optional(description),
        "blockers": normalize_optional(blockers),
        "efforts_hrs": efforts_hrs,
    }
    bucket.append(entry)
    save_tasks(data)
    return True, "Task saved successfully."


def list_projects(data: dict[str, Any]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for day_tasks in data.values():
        if not isinstance(day_tasks, list):
            continue
        for t in day_tasks:
            if not isinstance(t, dict):
                continue
            p = t.get("project")
            if not p:
                continue
            ps = str(p).strip()
            if not ps or ps == "-":
                continue
            if ps in seen:
                continue
            seen.add(ps)
            out.append(ps)
    out.sort(key=lambda s: s.lower())
    return out


def export_report_csv(path: str | Path, date_str: str, tasks: list[dict[str, Any]]) -> tuple[bool, str]:
    import csv

    p = Path(path)
    rows, total = tasks_table_rows(tasks)
    try:
        with p.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([f"Date: {date_str}"])
            w.writerow([])
            w.writerow(REPORT_HEADERS)
            w.writerows(rows)
            w.writerow([])
            w.writerow(["Total Efforts", format_hours(total)])
        return True, f"Exported CSV to {p}"
    except OSError as e:
        return False, f"Failed to export CSV: {e}"


def export_report_xlsx(path: str | Path, date_str: str, tasks: list[dict[str, Any]]) -> tuple[bool, str]:
    if openpyxl is None:
        return False, "Excel export needs openpyxl. Install it or export CSV."

    p = Path(path)
    rows, total = tasks_table_rows(tasks)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = date_str

        ws["A1"] = f"Date: {date_str}"
        if Font is not None:
            ws["A1"].font = Font(bold=True)

        start_row = 3
        for i, h in enumerate(REPORT_HEADERS, start=1):
            c = ws.cell(row=start_row, column=i, value=h)
            if Font is not None:
                c.font = Font(bold=True)

        for r_i, r in enumerate(rows, start=start_row + 1):
            for c_i, val in enumerate(r, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=val)
                if Alignment is not None and c_i in (4, 5):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        total_row = start_row + 1 + len(rows) + 2
        ws.cell(row=total_row, column=1, value="Total Efforts")
        ws.cell(row=total_row, column=2, value=float(total))

        for col in range(1, len(REPORT_HEADERS) + 1):
            max_len = 0
            for row in range(1, total_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max(10, max_len + 2), 60)

        wb.save(p)
        return True, f"Exported Excel to {p}"
    except OSError as e:
        return False, f"Failed to export Excel: {e}"

